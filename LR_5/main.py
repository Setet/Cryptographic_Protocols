import math
from datetime import datetime

import openpyxl as ox


def isPrime(n):
    if n == 2 or n == 3:
        return True
    elif n <= 1 or n % 2 == 0 or n % 3 == 0:
        return False

    # Чтобы проверить все числа вида 6k ± 1
    # пока i <= квадратный корень из n, с шагом 6
    for i in range(5, int(math.sqrt(n)) + 1, 6):
        if n % i == 0 or n % (i + 2) == 0:
            return False

    return True


if __name__ == '__main__':
    all_start_time = datetime.now()
    print("Я начал : "+str(all_start_time))
    n = 10000
    composite_numbers = []
    wb = ox.Workbook()
    ws = wb.worksheets[0]

    # Проверяем каждое число от 2 до n на простоту
    for i in range(2, n):

        if not isPrime(i):
            composite_numbers.append(i)  # Добавляем i в список составных чисел

    ws.cell(row=1, column=1).value = "Число"
    test_name = "Тест №"
    test_list = []

    for i in range(1, 101):
        count = test_name + str(i)
        test_list.append(count)

    for i, statN in enumerate(test_list):
        ws.cell(row=1, column=i + 2).value = statN

    for i, statN in enumerate(composite_numbers):
        ws.cell(row=i + 2, column=1).value = statN

    for i in range(2, len(composite_numbers) + 2):
        for j in range(2, len(test_list) + 2):
            ws.cell(row=i, column=j).value = str(isPrime(composite_numbers[i - 2]))

    wb.save('C:/Users/Setet/Documents/GitHub/Cryptographic_Protocols/LR_5_comparisons.xlsx')

    all_stop_time = datetime.now()
    print("Я закончил : "+str(all_stop_time))

    time_rap = all_stop_time-all_start_time
    print("Я работал : "+str(time_rap))
